"""
xlsx_parser.py — Parse TAX Invoice and Quotation xlsx files.

Cell layout verified against real samples (2026-05-18):

TAX INVOICE layout
  Row 9  : C12='No.:' C15=doc_number C21='Date/...' C25=date_raw
  Row 11 : C1='to:' C2=customer_name [+ TAX ID inline]
  Row 12 : C2=address line 1
  Row 13 : C2=address line 2
  Row 16 : header (No / Description / Quantity / Measure / Unit Price / Total)
  Row 17 : header (Thai)
  Row 18+ : items every 2 rows — C1=line_no C2=desc C17=qty C20=unit C23=price C26=total
  Row 47 : C10=label C24=subtotal (excl VAT)
  Row 48 : C10=label C24=vat
  Row 49 : C10=label C24=total (incl VAT)
  Row 51 : C1='Notes:' C3=notes line 1

QUOTATION layout
  Row 9  : C12='No.:' C15=doc_number C21='Date/...' C26=date_raw
  Row 11 : C1='to:' C2=customer_name (no TAX ID)
  Row 12 : C2=address line 1
  Row 13 : C2=address line 2
  Row 17 : header (No / Description / Quantity / Measure / Unit Price / Total)
  Row 18 : header (Thai)
  Row 19+ : items every 2 rows — C1=line_no C2=desc C17=qty C20=unit C24=price C28=total
  Row 43 : C10=label C25=subtotal (excl VAT)
  Row 44 : C10=label C25=vat
  Row 45 : C10=label C25=total (incl VAT)
  Row 47 : C1='หมายเหตุ:' C4=notes line 1
"""

import re
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from typing import Optional

import openpyxl


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------


@dataclass
class CustomerData:
    name: str
    tax_id: Optional[str]
    address: str


@dataclass
class ItemData:
    line_no: int
    description: str
    quantity: Decimal
    unit: str
    unit_price: Decimal
    total: Decimal


@dataclass
class DocumentData:
    doc_number: str
    doc_date_raw: object
    customer: CustomerData
    items: list
    subtotal: Decimal
    vat: Decimal
    total: Decimal
    notes: str


# ---------------------------------------------------------------------------
# Cell helpers
# ---------------------------------------------------------------------------


def _cell_str(ws, row: int, col: int) -> str:
    v = ws.cell(row=row, column=col).value
    if v is None:
        return ""
    return str(v).strip()


def _cell_decimal(ws, row: int, col: int) -> Decimal:
    v = ws.cell(row=row, column=col).value
    if v is None:
        return Decimal("0")
    try:
        return Decimal(str(v))
    except InvalidOperation:
        return Decimal("0")


def normalize_item_unit(description: str, unit: str) -> str:
    normalized = " ".join(unit.split())
    if normalized != "piece (บาน)":
        return normalized

    # "บาน" is correct for doors. Some sheet rows accidentally carry this
    # unit into timber/plank/decking lines; those should be counted as pieces.
    if "ประตู" in description:
        return normalized
    return "piece (แผ่น)"


def _split_customer(raw: str):
    """Extract name and TAX ID from a combined customer cell string."""
    for sep in ("TAX ID :", "TAX ID:", "เลขที่ผู้เสียภาษี"):
        if sep in raw:
            parts = raw.split(sep, 1)
            name = parts[0].strip()
            m = re.search(r"\d{13}", parts[1])
            tax_id = m.group(0) if m else parts[1].strip().split()[0]
            return name, tax_id
    return raw.strip(), None


def _collect_notes(ws, start_row: int, note_col: int, max_extra: int = 5) -> str:
    """Collect consecutive non-empty note lines from start_row downward."""
    lines = []
    for r in range(start_row, start_row + max_extra):
        v = _cell_str(ws, r, note_col)
        if v:
            lines.append(v)
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Public parsers
# ---------------------------------------------------------------------------


def parse_tax_invoice(filepath: str) -> DocumentData:
    """Parse a TAX Invoice xlsx file and return a DocumentData instance."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.worksheets[0]

    # Doc number and date — Row 9
    doc_number = _cell_str(ws, 9, 15)      # C15
    doc_date_raw = ws.cell(row=9, column=25).value  # C25

    # Customer — Row 11-13
    customer_raw = _cell_str(ws, 11, 2)    # C2
    name, tax_id = _split_customer(customer_raw)
    addr_parts = [_cell_str(ws, 12, 2), _cell_str(ws, 13, 2)]  # C2
    address = " ".join(p for p in addr_parts if p)

    # Detect layout variant: standard has EN header at row 16, variant at row 17
    if _cell_str(ws, 16, 2).lower() == 'description':
        item_start, item_stop, total_row = 18, 47, 47
    else:
        item_start, item_stop, total_row = 19, 48, 48

    # Items — every 2 rows from item_start until totals section
    items: list[ItemData] = []
    row = item_start
    while row < item_stop:
        line_no_val = ws.cell(row=row, column=1).value
        desc = _cell_str(ws, row, 2)
        if not desc and line_no_val is None:
            row += 2
            continue
        if not desc:
            row += 2
            continue
        qty = _cell_decimal(ws, row, 17)   # C17
        unit = normalize_item_unit(desc, _cell_str(ws, row, 20))      # C20
        price = _cell_decimal(ws, row, 23)  # C23
        total_cell = _cell_decimal(ws, row, 26)  # C26
        total = total_cell if total_cell else (qty * price).quantize(Decimal("0.01"))
        items.append(ItemData(
            line_no=len(items) + 1,
            description=desc,
            quantity=qty,
            unit=unit,
            unit_price=price,
            total=total,
        ))
        row += 2

    # Totals — col 24, rows total_row / +1 / +2
    subtotal = _cell_decimal(ws, total_row, 24)
    vat = _cell_decimal(ws, total_row + 1, 24)
    total_incl = _cell_decimal(ws, total_row + 2, 24)

    # Fallback: derive from items if cells are empty
    if total_incl == 0 and items:
        total_incl = sum(i.total for i in items)
    if subtotal == 0 and total_incl:
        subtotal = (total_incl / Decimal("107") * Decimal("100")).quantize(Decimal("0.01"))
    if vat == 0 and total_incl:
        vat = (total_incl / Decimal("107") * Decimal("7")).quantize(Decimal("0.01"))

    # Notes — Row 51, col 3 onward
    notes = _collect_notes(ws, 51, 3)

    return DocumentData(
        doc_number=doc_number,
        doc_date_raw=doc_date_raw,
        customer=CustomerData(name=name, tax_id=tax_id, address=address),
        items=items,
        subtotal=subtotal,
        vat=vat,
        total=total_incl,
        notes=notes,
    )


def parse_quotation(filepath: str) -> DocumentData:
    """Parse a Quotation xlsx file and return a DocumentData instance."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.worksheets[0]

    # Doc number and date — Row 9
    doc_number = _cell_str(ws, 9, 15)      # C15
    doc_date_raw = ws.cell(row=9, column=26).value  # C26

    # Customer — Row 11-13
    customer_raw = _cell_str(ws, 11, 2)    # C2
    name, tax_id = _split_customer(customer_raw)
    addr_parts = [_cell_str(ws, 12, 2), _cell_str(ws, 13, 2)]  # C2
    address = " ".join(p for p in addr_parts if p)

    # Items — start row 19, every 2 rows, until totals section
    # Quotation totals start at row 43
    items: list[ItemData] = []
    row = 19
    while row < 43:
        line_no_val = ws.cell(row=row, column=1).value
        desc = _cell_str(ws, row, 2)
        if not desc and line_no_val is None:
            row += 2
            continue
        if not desc:
            row += 2
            continue
        qty = _cell_decimal(ws, row, 17)   # C17
        unit = normalize_item_unit(desc, _cell_str(ws, row, 20))      # C20
        price = _cell_decimal(ws, row, 24)  # C24 (differs from TI)
        total_cell = _cell_decimal(ws, row, 28)  # C28 (differs from TI)
        total = total_cell if total_cell else (qty * price).quantize(Decimal("0.01"))
        items.append(ItemData(
            line_no=len(items) + 1,
            description=desc,
            quantity=qty,
            unit=unit,
            unit_price=price,
            total=total,
        ))
        row += 2

    # Totals — Rows 43-45, col 25
    subtotal = _cell_decimal(ws, 43, 25)   # C25 row 43
    vat = _cell_decimal(ws, 44, 25)        # C25 row 44
    total_incl = _cell_decimal(ws, 45, 25)  # C25 row 45

    # Fallback: derive from items if cells are empty
    if total_incl == 0 and items:
        total_incl = sum(i.total for i in items)
    if subtotal == 0 and total_incl:
        subtotal = (total_incl / Decimal("107") * Decimal("100")).quantize(Decimal("0.01"))
    if vat == 0 and total_incl:
        vat = (total_incl / Decimal("107") * Decimal("7")).quantize(Decimal("0.01"))

    # Notes — Row 47, col 4 onward
    notes = _collect_notes(ws, 47, 4)

    return DocumentData(
        doc_number=doc_number,
        doc_date_raw=doc_date_raw,
        customer=CustomerData(name=name, tax_id=tax_id, address=address),
        items=items,
        subtotal=subtotal,
        vat=vat,
        total=total_incl,
        notes=notes,
    )
