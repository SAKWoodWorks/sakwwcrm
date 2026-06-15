from pathlib import Path

import openpyxl

from import_products import read_products


def test_read_products_finds_teak_dashboard_code_and_full_name(tmp_path: Path):
    workbook_path = tmp_path / "teak-stock.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Teak DASHBOARD"
    ws.cell(12, 1).value = "CODE"
    ws.cell(12, 30).value = "Full Name"
    ws.cell(17, 1).value = "TPARQ1590450"
    ws.cell(17, 30).value = "Teak Parquet Boards"
    wb.save(workbook_path)

    products = read_products(workbook_path, "Teak DASHBOARD")

    assert products == [
        {
            "sku_code": "TPARQ1590450",
            "full_name": "Teak Parquet Boards",
            "thickness": None,
            "width": None,
            "length": None,
            "volume": None,
            "weight": None,
            "rt_cost": None,
            "ws_cost": None,
        }
    ]


def test_read_products_uses_teak_grade_ab_retail_as_rt_cost(tmp_path: Path):
    workbook_path = tmp_path / "teak-stock.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Teak DASHBOARD"
    ws.cell(12, 1).value = "CODE"
    ws.cell(12, 28).value = "CBM/ 1 PCS"
    ws.cell(12, 29).value = "KG/ 1 PCS"
    ws.cell(12, 30).value = "Full Name"
    ws.cell(14, 25).value = 'Grade "AB" Retail'
    ws.cell(17, 1).value = "TPARQ1590450"
    ws.cell(17, 25).value = 60.35
    ws.cell(17, 28).value = 0.0006075
    ws.cell(17, 29).value = 0.3888
    ws.cell(17, 30).value = "Teak Parquet Boards"
    wb.save(workbook_path)

    products = read_products(workbook_path, "Teak DASHBOARD")

    assert str(products[0]["rt_cost"]) == "60.35"


def test_read_products_parses_dimensions_from_teak_full_name(tmp_path: Path):
    workbook_path = tmp_path / "teak-stock.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Teak DASHBOARD"
    ws.cell(12, 1).value = "CODE"
    ws.cell(12, 30).value = "Full Name"
    ws.cell(17, 1).value = "TPARQ1590450"
    ws.cell(17, 30).value = "Teak Parquet Boards ไม้ปาร์เก้สัก (15x90x450 มม./mm.)"
    wb.save(workbook_path)

    products = read_products(workbook_path, "Teak DASHBOARD")

    assert str(products[0]["thickness"]) == "15"
    assert str(products[0]["width"]) == "90"
    assert str(products[0]["length"]) == "0.45"
