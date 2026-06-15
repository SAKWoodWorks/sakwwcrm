"""
Import products from docs/Stock.xlsx DASHBOARD sheet into PostgreSQL.
Usage: python import_products.py [--dry-run]
"""
import argparse
import re
import sys
from decimal import Decimal, InvalidOperation
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

from dotenv import load_dotenv
load_dotenv(Path(__file__).parent.parent / ".env.local")

import os
import openpyxl
import psycopg2

DEFAULT_STOCK_FILE = Path(__file__).parent.parent / "docs" / "Stock.xlsx"
DEFAULT_SHEET = "DASHBOARD"
HEADER_SCAN_ROWS = 40
HEADER_SCAN_COLS = 80


def _dec(v) -> Decimal | None:
    if v is None:
        return None
    try:
        return Decimal(str(v))
    except InvalidOperation:
        return None


def _parse_dimensions(full_name: str):
    match = re.search(
        r"\((\d+(?:\.\d+)?)\s*[x×]\s*(\d+(?:\.\d+)?)\s*[x×]\s*(\d+(?:\.\d+)?)\s*(?:มม|mm)",
        full_name,
        re.IGNORECASE,
    )
    if not match:
        return None, None, None
    thickness = _dec(match.group(1))
    width = _dec(match.group(2))
    length = _dec(Decimal(match.group(3)) / Decimal("1000"))
    return thickness, width, length


def _norm_header(v):
    return str(v).strip().lower() if v is not None else ""


def _find_header_columns(ws):
    for r in range(1, min(ws.max_row, HEADER_SCAN_ROWS) + 1):
        code_col = full_name_col = volume_col = weight_col = None
        for c in range(1, min(ws.max_column, HEADER_SCAN_COLS) + 1):
            header = _norm_header(ws.cell(r, c).value)
            if header == "code":
                code_col = c
            elif header == "full name":
                full_name_col = c
            elif header == "cbm/ 1 pcs":
                volume_col = c
            elif header == "kg/ 1 pcs":
                weight_col = c

        if code_col and full_name_col:
            return {
                "header_row": r,
                "data_start_row": r + 5,
                "code_col": code_col,
                "full_name_col": full_name_col,
                "volume_col": volume_col,
                "weight_col": weight_col,
                "rt_cost_col": _find_first_header_col(ws, ['grade "a" retail', 'grade "ab" retail'], r, r + 7),
                "ws_cost_col": _find_header_col(ws, 'grade "a" wholesale', r, r + 7),
            }

    raise ValueError("Could not find CODE and Full Name columns in dashboard sheet")


def _find_header_col(ws, target, start_row, end_row):
    for r in range(start_row, min(ws.max_row, end_row) + 1):
        for c in range(1, min(ws.max_column, HEADER_SCAN_COLS) + 1):
            if _norm_header(ws.cell(r, c).value) == target:
                return c
    return None


def _find_first_header_col(ws, targets, start_row, end_row):
    for target in targets:
        col = _find_header_col(ws, target, start_row, end_row)
        if col:
            return col
    return None


def read_products(stock_file=DEFAULT_STOCK_FILE, sheet_name=DEFAULT_SHEET):
    wb = openpyxl.load_workbook(str(stock_file), data_only=True)
    ws = wb[sheet_name]
    columns = _find_header_columns(ws)
    products = []
    for r in range(columns["data_start_row"], ws.max_row + 1):
        sku_code = ws.cell(r, columns["code_col"]).value
        if not sku_code or not isinstance(sku_code, str):
            continue
        full_name = ws.cell(r, columns["full_name_col"]).value or sku_code
        thickness, width, length = _parse_dimensions(str(full_name))
        volume = _dec(ws.cell(r, columns["volume_col"]).value) if columns["volume_col"] else None
        weight = _dec(ws.cell(r, columns["weight_col"]).value) if columns["weight_col"] else None
        rt_cost = _dec(ws.cell(r, columns["rt_cost_col"]).value) if columns["rt_cost_col"] else None
        ws_cost = _dec(ws.cell(r, columns["ws_cost_col"]).value) if columns["ws_cost_col"] else None
        products.append({
            "sku_code": sku_code.strip(),
            "full_name": str(full_name).strip(),
            "thickness": thickness,
            "width": width,
            "length": length,
            "volume": volume,
            "weight": weight,
            "rt_cost": rt_cost,
            "ws_cost": ws_cost,
        })
    return products


def upsert_products(products, dry_run=False):
    db_url = os.environ["DATABASE_URL"]
    conn = psycopg2.connect(db_url)
    conn.autocommit = False

    inserted = updated = skipped = 0
    try:
        with conn.cursor() as cur:
            for p in products:
                if dry_run:
                    print(f"  [{p['sku_code']}] {p['full_name'][:50]}  ws={p['ws_cost']}  rt={p['rt_cost']}")
                    continue

                cur.execute("""
                    INSERT INTO products (
                        sku_code, full_name, thickness, width, length,
                        volume, weight, ws_cost, rt_cost, updated_at
                    )
                    VALUES (
                        %(sku_code)s, %(full_name)s, %(thickness)s, %(width)s, %(length)s,
                        %(volume)s, %(weight)s, %(ws_cost)s, %(rt_cost)s, NOW()
                    )
                    ON CONFLICT (sku_code) DO UPDATE SET
                        full_name  = EXCLUDED.full_name,
                        thickness  = COALESCE(EXCLUDED.thickness, products.thickness),
                        width      = COALESCE(EXCLUDED.width, products.width),
                        length     = COALESCE(EXCLUDED.length, products.length),
                        volume     = COALESCE(EXCLUDED.volume, products.volume),
                        weight     = COALESCE(EXCLUDED.weight, products.weight),
                        ws_cost    = COALESCE(EXCLUDED.ws_cost, products.ws_cost),
                        rt_cost    = COALESCE(EXCLUDED.rt_cost, products.rt_cost),
                        updated_at = NOW()
                    RETURNING (xmax = 0) AS inserted
                """, p)
                row = cur.fetchone()
                if row and row[0]:
                    inserted += 1
                else:
                    updated += 1

        if not dry_run:
            conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    return inserted, updated


def main():
    import io

    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--file", type=Path, default=DEFAULT_STOCK_FILE)
    parser.add_argument("--sheet", default=DEFAULT_SHEET)
    args = parser.parse_args()

    print(f"Reading {args.file.name} {args.sheet} sheet...")
    products = read_products(args.file, args.sheet)
    print(f"Found {len(products)} products")

    if args.dry_run:
        print("\n--- Dry run ---")
        for p in products:
            print(f"  [{p['sku_code']}] {p['full_name'][:60]}  vol={p['volume']}  wt={p['weight']}  ws={p['ws_cost']}  rt={p['rt_cost']}")
        return

    inserted, updated = upsert_products(products)
    print(f"Done: {inserted} inserted, {updated} updated")


if __name__ == "__main__":
    main()
